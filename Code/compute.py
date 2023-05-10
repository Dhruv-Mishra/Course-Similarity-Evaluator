import pandas as pd
import os
import glob
import openpyxl
import warnings
import re
import tensorflow_hub as hub
import difflib
import numpy as np
import math
from sklearn.metrics.pairwise import linear_kernel
import pickle
import nltk
nltk.download('punkt',quiet=True)
nltk.download("stopwords",quiet=True)
import string 
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords

warnings.simplefilter(action='ignore', category=UserWarning)

embed = hub.load("https://tfhub.dev/google/universal-sentence-encoder/4")

class Course_Loader:
    course_names = {}
    course_codes = {}
    names_list = []
    codes_list = []
    df = -1
    newEntries = False
    cosine_sim = -1
    indices = -1

    def __init__(self):
        data = ['Course Name','Course Code','Course Description']
        self.df = pd.DataFrame(columns = data)
        self.course_codes = {}
        self.course_names = {}
        self.names_list = []
        self.codes_list = []
        self.newEntries = False

    def getDataset(self):
        return self.df
    
    def addCourse(self, f):
        newCourse = Course(f)
        self.course_names[newCourse.name] = newCourse
        self.course_codes[newCourse.code] = newCourse
        new_df = pd.DataFrame({"Course Name":[newCourse.name],"Course Code":[newCourse.code], "Course Description":[newCourse.description]})
        self.df = self.df.append(new_df,ignore_index=True)
        self.newEntries = True # Keeps track of if new entries were added after calculating the similarity check 

    def findbyCode(self,code, exact = False):
        ans = []
        if(not exact):
            code = "".join(code.lower().split())
            out = difflib.get_close_matches(code, self.codes_list, cutoff= 0.0000001, n = 10)
            for i in out:
                ans.append(self.course_codes[i])
        if(code in self.course_codes.keys()):
            ans.append(self.course_codes[code])
        return ans
    
    def findbyName(self,name, exact = False):
        ans = []
        if(not exact):
            name = "".join(name.lower().split())
            out = difflib.get_close_matches(name, self.names_list, cutoff= 0.0000001, n = 10)
            for i in out:
                ans.append(self.course_names[i])
        if(name in self.course_names.keys()):
            ans.append(self.course_names[name])
        return ans

    def initialize_recommender(self):
        if(self.newEntries):
            self.newEntries = False
            self.df = self.df.dropna()
            self.df = self.df.drop_duplicates(subset='Course Name', keep='first')
            self.df = self.df.reset_index(drop=True)
            self.names_list = self.df['Course Name'].unique()
            self.codes_list = self.df['Course Code'].unique()
            encodings = embed(self.df['Course Description'])
            matrix = np.vstack(encodings)
            cosine_sim = linear_kernel(matrix, matrix)
            indices = pd.Series(self.df.index, index=self.df['Course Name'])
            self.cosine_sim = cosine_sim
            self.indices = indices

    def get_recommendations(self, title):
        idx = self.indices[title]
        sim_scores = list(enumerate(self.cosine_sim[idx]))
        return sim_scores
    
    def combine_recommendations(self,titles):
        out = []
        self.initialize_recommender()
        for i in range(len(self.df)):
            out.append([i,0])
        for org_title in titles:            #### Names may or may not be present in the dataset, finds the closest title matching in the dataset
            title = difflib.get_close_matches(org_title, self.names_list, cutoff= 0.0000001, n = 1)[0]  #Searching the closest match in the dataset 
            cur_out = self.get_recommendations(title)
            for score in range(len(cur_out)):
                out[score][1] += cur_out[score][1]
        final_score = sorted(out, key=lambda x: x[1], reverse=True)
        df_indices = [i[0] for i in final_score][1:10]
        result = self.df.iloc[df_indices]
        return result
                     
class Course:

    course_name_list = set([])
    file_name = ""
    code = ""
    description = ""
    embedding = []
    similarity_matrix = []
    file_ptr = ""
    name = ""

    def pre_process_text(self, new_s):
        new_s = ".".join(new_s.split("\n"))
        #new_s = " ".join(new_s.split("-"))
        new_s = new_s.lower()
        # translate_table = dict((ord(char), " ") for char in string.punctuation)   
        # new_s = new_s.translate(translate_table)
        # li = word_tokenize(new_s)
        # stop_words = set(stopwords.words("english"))
        # filter_li = []
        # for words in li:
        #     if(words not in stop_words):
        #         filter_li.append(words)
        # ans = " ".join(filter_li)
        ans = new_s
        return ans

    def __init__(self, file_pointer):
        self.file_name = str(f)
        self.file_ptr = file_pointer
        self.description = self.find_description() + " " + self.find_course_plan()
        self.code = self.find_code()
        self.name = self.find_name()


    def find_description(self):
        workbook = openpyxl.load_workbook(self.file_ptr)
        worksheet = workbook.active
        for coll in range(1,6):
            for roww in range(1,20):
                if("description" in "".join(re.sub(r'[^\w\s]','', re.sub(r'\d+','',str(worksheet.cell(row=roww, column=coll).value).lower())).split())):
                    return str(worksheet.cell(row=roww, column=coll+1).value)
        return "-1"
    
    def find_code(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('code' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name 
    
    def find_name(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('name' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name
    
    def find_course_plan(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        lec_cell = (-1,-1)
        for coll in range(1,20):
            for roww in range(1,41):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if(len(cell_val) >= 2 and len(cell_val) <= 30 and 'lecture' in cell_val and 'topic' in cell_val):
                    lec_cell = (roww,coll)
                    break
            if(lec_cell[0]!=-1):
                break
        if(lec_cell[0] == -1):
            return "-1"
        plan = ""
        for i in range(8):
            new_row = lec_cell[0]+i+1
            new_col = lec_cell[1]
            cell_val = str(worksheet.cell(row=new_row, column=new_col).value).lower().strip()
            plan = plan + ". " + cell_val
        plan = self.pre_process_text(plan)
        plan = ". ".join(plan.split("•"))
        plan = plan + "."
        return plan
    
    def compute_embedding(self):
        return embed([self.description])
   
try:                                              #Loading Course_Loader object if it exists
    #print(1/0)
    with open("Course_Loader_Save", "rb") as f:
        c = pickle.load(f)
except:                                           #Creating a new object if it doesn't exist and saving it on the disk

    path = os.getcwd() # Getting the current directory to access the data files
    path = str(path) + "\\Data\\" # Data must be stored inside the Data folder located in the current directory 

    csv_files = glob.glob(os.path.join(path, "*.xlsx")) # Get names of all csv files
    
    c = Course_Loader()

    for f in csv_files:  #Adding files to the course loader object 
        c.addCourse(f)

    with open("Course_Loader_Save", "wb") as f:   #Saving the new object on the disk 
        pickle.dump(c,f)

query = ["machine learning"]
c.combine_recommendations(query) #Making a query
