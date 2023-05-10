import pandas as pd
import os
import glob
import openpyxl
import warnings
import re
import tensorflow_hub as hub
import difflib
import numpy as np
import math
from sklearn.metrics.pairwise import linear_kernel
import pickle

warnings.simplefilter(action='ignore', category=UserWarning)

embed = hub.load("https://tfhub.dev/google/universal-sentence-encoder/4")

class Course:

    course_name_list = set([])
    file_name = ""
    code = ""
    description = ""
    embedding = []
    similarity_matrix = []
    file_ptr = ""
    name = ""

    def pre_process_text(self, new_s):
        new_s = ".".join(new_s.split("\n"))
        #new_s = " ".join(new_s.split("-"))
        new_s = new_s.lower()
        # translate_table = dict((ord(char), " ") for char in string.punctuation)   
        # new_s = new_s.translate(translate_table)
        # li = word_tokenize(new_s)
        # stop_words = set(stopwords.words("english"))
        # filter_li = []
        # for words in li:
        #     if(words not in stop_words):
        #         filter_li.append(words)
        # ans = " ".join(filter_li)
        ans = new_s
        return ans

    def __init__(self, file_pointer):
        self.file_name = str(f)
        self.file_ptr = file_pointer
        self.description = self.find_description() + " " + self.find_course_plan()
        self.code = self.find_code()
        self.name = self.find_name()


    def find_description(self):
        workbook = openpyxl.load_workbook(self.file_ptr)
        worksheet = workbook.active
        for coll in range(1,6):
            for roww in range(1,20):
                if("description" in "".join(re.sub(r'[^\w\s]','', re.sub(r'\d+','',str(worksheet.cell(row=roww, column=coll).value).lower())).split())):
                    return str(worksheet.cell(row=roww, column=coll+1).value)
        return "-1"
    
    def find_code(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('code' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name 
    
    def find_name(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('name' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name
    
    def find_course_plan(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        lec_cell = (-1,-1)
        for coll in range(1,20):
            for roww in range(1,41):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if(len(cell_val) >= 2 and len(cell_val) <= 30 and 'lecture' in cell_val and 'topic' in cell_val):
                    lec_cell = (roww,coll)
                    break
            if(lec_cell[0]!=-1):
                break
        if(lec_cell[0] == -1):
            return "-1"
        plan = ""
        for i in range(8):
            new_row = lec_cell[0]+i+1
            new_col = lec_cell[1]
            cell_val = str(worksheet.cell(row=new_row, column=new_col).value).lower().strip()
            plan = plan + ". " + cell_val
        plan = self.pre_process_text(plan)
        plan = ". ".join(plan.split("•"))
        plan = plan + "."
        return plan
    
    def compute_embedding(self):
        return embed([self.description])
