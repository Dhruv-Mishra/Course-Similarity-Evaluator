import pandas as pd
import os
import glob
import openpyxl
import warnings
import re
import tensorflow_hub as hub
import difflib
import numpy as np
import math
from sklearn.metrics.pairwise import linear_kernel
import pickle

warnings.simplefilter(action='ignore', category=UserWarning)
class Course:
    course_name_list = set([])
    file_name = ""
    code = ""
    description = ""
    embedding = []
    similarity_matrix = []
    file_ptr = ""
    name = ""

    def __init__(self, file_pointer):
        self.file_name = str(f)
        self.file_ptr = file_pointer
        self.description = self.find_description()
        self.code = self.find_code()
        self.name = self.find_name()
        if(self.description != -1):
            self.embedding = self.compute_embedding()

    def find_description(self):
        workbook = openpyxl.load_workbook(self.file_ptr)
        worksheet = workbook.active
        description_row = -1
        description_col = -1
        for coll in range(1,6):
            for roww in range(1,20):
                if("description" in "".join(re.sub(r'[^\w\s]','', re.sub(r'\d+','',str(worksheet.cell(row=roww, column=coll).value).lower())).split())):
                    description_row = roww
                    description_col = coll
                    return str(worksheet.cell(row=roww, column=coll+1).value)
        return "-1"
    
    def find_code(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('code' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name 
    
    def find_name(self):
        f = self.file_ptr
        workbook = openpyxl.load_workbook(f)
        worksheet = workbook.active
        for coll in range(1,4):
            for roww in range(1,5):
                cell_val = "".join("".join(str(worksheet.cell(row=roww, column=coll).value).lower().strip().split()).split("-"))
                if('name' in cell_val):
                    cell_val = " ".join("".join(str(worksheet.cell(row=roww, column=coll+1).value).lower().strip().split()).split("-"))
                    return cell_val
        #If not found in the sheet, find in file_name 
        file_name = str(f).split("\\")[-1]
        file_name = file_name.split("/")[-1]
        file_name = file_name.split(".")
        if(len(file_name) >= 2):
            file_name = file_name[-2]
        file_name = "".join(file_name.lower().split())
        return file_name

    def compute_embedding(self):
        return embed([self.description])